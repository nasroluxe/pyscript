
import openpyxl
from bs4 import BeautifulSoup

workbook = openpyxl.load_workbook('text.xlsx')
worksheet = workbook.active

with open('rewriteA.txt', 'w', encoding='utf-8') as file:
    for row in range(2, worksheet.max_row+1):
        cell_a_value = worksheet.cell(row=row, column=1).value
        cell_c_value = worksheet.cell(row=row, column=3).value


        replaced_link_text = {}
        for link in links:
            link_text = link.string
            hyperlink_tag = f'<a href="{link["href"]}"'
            if link.has_attr('rel'):
                hyperlink_tag += f' rel="{link["rel"][0]}"'
                if len(link['rel']) > 1:
                    hyperlink_tag += f' {link["rel"][1]}'
            if link.has_attr('target'):
                hyperlink_tag += f' target="{link["target"]}"'
            hyperlink_tag += f'>{link_text}</a>'

            if link_text in replaced_link_text:
                last_index = cell_c_value[::-1].index(link_text[::-1])
                cell_c_value = cell_c_value[::-1][:last_index][::-1] + hyperlink_tag[::-1] + cell_c_value[::-1][last_index+len(link_text):][::-1]
            else:
                last_index = cell_c_value.rindex(link_text)
                cell_c_value = cell_c_value[:last_index] + hyperlink_tag + cell_c_value[last_index+len(link_text):]
            replaced_link_text[link_text] = True

        soup = BeautifulSoup(cell_a_value, 'html.parser')
        strongs = soup.find_all('strong')

        replaced_strong_text = {}
        for strong in strongs:
            strong_text = strong.string
            strong_tag = f'<strong>{strong_text}</strong>'

            if strong_text in cell_c_value:
                if strong_text in replaced_strong_text:
                    last_index = cell_c_value[::-1].index(strong_text[::-1])
                    cell_c_value = cell_c_value[::-1][:last_index][::-1] + strong_tag[::-1] + cell_c_value[::-1][last_index+len(strong_text):][::-1]
                else:
                    last_index = cell_c_value.rindex(strong_text)
                    cell_c_value = cell_c_value[:last_index] + strong_tag + cell_c_value[last_index+len(strong_text):]
                replaced_strong_text[strong_text] = True

        worksheet.cell(row=row, column=4).value = cell_c_value

        soup = BeautifulSoup(cell_a_value, 'html.parser')
        text_only = ' '.join(soup.stripped_strings)
        worksheet.cell(row=row, column=2).value = text_only

workbook.save('result.xlsx')

with open('rewriteA.txt', 'w', encoding='utf-8') as file:
    for row in range(2, worksheet.max_row+1):
        cell_d_value = worksheet.cell(row=row, column=4).value
        file.write(cell_d_value + '\n')

print('Done')
